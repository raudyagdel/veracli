import os
import argparse
import sys
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from datetime import datetime

RISK_RATING = {
        "0": {"name": "Unassessable", "icon": "img/unassessable.png"},
        "2": {"name": "Low", "icon": "img/low.png"},
        "3": {"name": "Medium", "icon": "img/medium.png"},
        "4": {"name": "High", "icon": "img/high.png"},
        }

font_path = "font/SoleilRegular.ttf"

# Define an Enum for risk_rating
class RiskRating:
    
    def from_value(risk_rating_value):
        """
        Get the risk rating details (name and icon) based on the numeric value.
        """
        return RISK_RATING.get(risk_rating_value, {"name": "Unknown", "icon": None})

    def to_string(risk_rating_value):
        """
        Convert the risk rating value to a human-readable name.
        """
        return RISK_RATING.get(risk_rating_value, {"name": "Unknown", "icon": None})["name"]
    
    def get_icon(risk_rating_value):
        """
        Get the icon path or image associated with the risk rating value.
        """
        # Try to get the icon from the dictionary, or return None if it's not found
        icon_info = RISK_RATING.get(risk_rating_value)
        if icon_info and "icon" in icon_info:
            return icon_info["icon"]  # Return the icon path or URL
        else:
            return None  # Return None if no icon is associated

def extract_metada(xml_file): 
    # Create element tree object
    tree = ET.parse(xml_file)

    # Get the root element (which should be 'detailedreport' with a namespace)
    root = tree.getroot()

    # Check if we have the correct root tag with the namespace
    if root.tag != '{https://www.veracode.com/schema/reports/export/1.0}detailedreport':
        print("Root element is not 'detailedreport'.")
        return

    # Extract the required attributes
    metadata = {
                "app_name": root.attrib.get('app_name', None),
                "sandbox_name": root.attrib.get('sandbox_name', None),
                "version": root.attrib.get('version', None),
            }
    
    return metadata
    
def generate_filename(metadata):
    # Replace blank spaces with underscores for each metadata field
    app_name = metadata.get("app_name", "").replace(" ", "_")
    sandbox_name = metadata.get("sandbox_name", "").replace(" ", "_")
    version = metadata.get("version", "").replace(" ", "_")
    
    # Get the current timestamp in the format 'YYYY-MM-DD_HH-MM-SS'
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    
    # Construct the filename
    filename = f"licensereport_{app_name}_{sandbox_name}_{version}.xlsx"
    
    return filename
        
def extract_license(xml_file): 
    # Create element tree object
    tree = ET.parse(xml_file)

    # Get the root element (which should be 'detailedreport' with a namespace)
    root = tree.getroot()

    # Define the namespace dictionary
    namespaces = {
        '': 'https://www.veracode.com/schema/reports/export/1.0',  # Default namespace
    }

    # Check if we have the correct root tag with the namespace
    if root.tag != '{https://www.veracode.com/schema/reports/export/1.0}detailedreport':
        print("Root element is not 'detailedreport'.")
        return

    # Extract the required attributes
    app_name = root.attrib.get('app_name', 'Unknown')
    sandbox_name = root.attrib.get('sandbox_name', 'Unknown')
    version = root.attrib.get('version', 'Unknown')
    print("﹡﹡﹡﹡﹡﹡﹡﹡﹡﹡﹡")
    print("﹡  LICENSE REPORT  ﹡")
    print("﹡﹡﹡﹡﹡﹡﹡﹡﹡﹡﹡")
    print(f"App Name: {app_name}")
    print(f"Sandbox Name: {sandbox_name}")
    print(f"Version: {version}")
    
    # Create empty list for components
    components = []

    # Find the 'software_composition_analysis' -> 'vulnerable_components' -> 'component' elements
    software_composition_analysis = root.find(".//{https://www.veracode.com/schema/reports/export/1.0}software_composition_analysis", namespaces)
    if software_composition_analysis is not None:
        vulnerable_components = software_composition_analysis.find("{https://www.veracode.com/schema/reports/export/1.0}vulnerable_components", namespaces)
        if vulnerable_components is not None:
            # Iterate through each 'component' inside 'vulnerable_components'
            for component in vulnerable_components.findall("{https://www.veracode.com/schema/reports/export/1.0}component", namespaces):
                # For each component, get license details
                licenses = component.find("{https://www.veracode.com/schema/reports/export/1.0}licenses", namespaces)
                if licenses is not None:
                    for license in licenses.findall("{https://www.veracode.com/schema/reports/export/1.0}license", namespaces):
                        # Extract license info from each license tag
                        risk_rating_value = str(license.attrib.get('risk_rating'))
                        
                        # Extract license info from each license tag
                        license_info = {
                            "file_name": component.attrib.get('file_name'),
                            "license_name": license.attrib.get('name'),
                            "spdx_id": license.attrib.get('spdx_id'),
                            "license_url": license.attrib.get('license_url'),
                            "risk_rating": risk_rating_value
                        }
                        components.append(license_info)  # Add license info to the list

    # # Sort the components list by risk_rating in descending order
    # components.sort(key=lambda x: int(x['risk_rating']) if x['risk_rating'].isdigit() else -1, reverse=True)
    
     # Sort the components list first by risk_rating (descending), then by license_name (ascending)
    components.sort(key=lambda x: (int(x['risk_rating']) if x['risk_rating'].isdigit() else -1, x['license_name'] or ''))  # Second criterion: license_name, empty string if None
    # Sort in descending order for risk_rating, ascending order for license_name
    components.sort(key=lambda x: int(x['risk_rating']) if x['risk_rating'].isdigit() else -1, reverse=True)

    # Check if we found any components with license info and print them
    return components


def export_to_excel(components, output_file='licenses.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Licenses'

    # Load the custom font
    if font_path:
        custom_font = Font(name="Soleil")
    else:
        custom_font = Font()
        
    # Define header row
    headers = ['Component Filename', 'License', 'License Risk',]
    ws.append(headers)

    # Make the header row bold
    for cell in ws[1]:  # Access the first row (header row)
        cell.font = Font(bold=True)
    
    # Style for odd and even rows
    odd_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")  
    even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  #
    
     # Define row height
    row_height = 25
    
    # Add data rows and apply styles
    for idx, component in enumerate(components, start=1):
        row = [
            component['file_name'],
            component['license_name'],
            # component['spdx_id'],
            # component['license_url']
        ]
        
        ws.append(row)
        
        # Apply row height
        ws.row_dimensions[idx + 1].height = row_height 
        
        # Style the rows (odd/even)
        row_fill = odd_fill if idx % 2 != 0 else even_fill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=idx + 1, column=col)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = custom_font

        # Add hyperlink to the 'License' column
        license_name = component['license_name']
        license_url = component['license_url']
        
        if license_url:
            cell = ws.cell(row=idx + 1, column=2)  # 'License' column
            cell.hyperlink = license_url
            cell.font = Font(color="4E9EBF", underline="single")  # Blue, underlined font for hyperlink
        
        # # Change font color for 'file_name' in the first column
        # file_name_cell = ws.cell(row=idx + 1, column=1)  # 'File Name' column
        # file_name_cell.font = Font(color="FF0000")  # Red font color for the file name
        
        # Add icons for risk_rating
        risk_rating = component['risk_rating']
        risk_rating_icon_path = RiskRating.get_icon(risk_rating)  # Get the icon path
        risk_rating_risk_name = RiskRating.to_string(risk_rating)  # Get the risk rating text
        
        try:
            if risk_rating_icon_path:
                if os.path.exists(risk_rating_icon_path):
                    img = Image(risk_rating_icon_path)
                    img.width = 25
                    img.height = 25
                    
                    # Merge cells E and F to display both image and text in the same area
                    ws.merge_cells(start_row=idx + 1, start_column=5, end_row=idx + 1, end_column=6)
                    icon_cell = f'C{idx + 1}'  # Image will be placed in column E
                    
                    # Adjust the vertical centering by calculating the row offset
                    offset_y = (row_height - img.height) // 2  # Center the image vertically in the row
                    
                    img1 = openpyxl.drawing.image.Image(risk_rating_icon_path)
                    p2e = pixels_to_EMU
                    h, w = img1.height, img1.width
                    position = XDRPoint2D(p2e(500), p2e(500))
                    size = XDRPositiveSize2D(p2e(h), p2e(w))

                    img1.anchor = AbsoluteAnchor(pos=position, ext=size)
                    # ws.add_image(img1)

                    ws.add_image(img1, icon_cell)  # Add image over cell in column 'E'

                    # Add the text (Risk Rating) in the merged cell, aligned next to the image
                    text_cell = f'C{idx + 1}'
                    ws[text_cell] = "    " + risk_rating_risk_name  # Add text to merged cell
                    ws[text_cell].alignment = Alignment(horizontal="left", vertical="center")  # Align text to the left
        except Exception as e:
            print("Error adding icon:", e)

    
    # Adjust column widths based on the data length
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Add some padding for readability
        ws.column_dimensions[column].width = adjusted_width
        
    # Save the file
    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(description="Extract software composicion licenses and convert output to xlxs.")
    parser.add_argument("--xml", help="xml file to extract license")
 
    args = parser.parse_args()

    # Check if essential arguments are missing and show help if true
    if not args.xml:
        print("Error: Missing required arguments!")
        parser.print_help()
        sys.exit(1) 
    
    # Extract License
    licences = extract_license(args.xml)
    
    # Get App MetaData Info 
    metadata = extract_metada(args.xml)
    output_file_name = generate_filename(metadata)
    export_to_excel(licences, output_file_name)

if __name__ == "__main__":
    main()